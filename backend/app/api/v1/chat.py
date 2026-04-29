"""
Chat endpoints — drive the LangGraph workflow.

POST /chat/start          — start a new thread for a session (kicks off ingest → generate)
POST /chat/{thread_id}/message  — send a message (during clarification or free chat)
POST /chat/{thread_id}/review   — submit human review decision (approve / feedback)
GET  /chat/{thread_id}/status   — poll for current step + latest message
GET  /chat/{thread_id}/test-cases — retrieve generated test cases
POST /chat/{thread_id}/execute  — trigger test execution
"""
from __future__ import annotations

import asyncio
import json
import os
import re
import shutil
import uuid
from io import BytesIO
from pathlib import Path

import openpyxl
import openpyxl.utils
import subprocess
from openpyxl.styles import Alignment, Font, PatternFill

from fastapi import APIRouter, BackgroundTasks, Depends, HTTPException
from fastapi.responses import StreamingResponse
from langchain_core.messages import AIMessage, HumanMessage
from pydantic import BaseModel

from app.api.deps import get_graph
from app.api.v1.report import generate_allure_report
from app.core.config import settings
from app.core.logging import logger
from app.services.session_registry import registry
from app.models.schemas import (
    ChatRequest,
    HumanReviewRequest,
    QuestionnaireSubmitRequest,
    ResumeResponse,
    SessionStatus,
    SessionStatusResponse,
)

router = APIRouter()


def _config(thread_id: str) -> dict:
    return {"configurable": {"thread_id": thread_id}}


def _last_ai_message(state: dict) -> str:
    messages = state.get("messages", [])
    for msg in reversed(messages):
        if hasattr(msg, "content") and msg.__class__.__name__ == "AIMessage":
            return msg.content
    return ""


def _map_status(step: str | None) -> SessionStatus:
    mapping = {
        "parsing": SessionStatus.PARSING,
        "awaiting_questionnaire": SessionStatus.AWAITING_QUESTIONNAIRE,
        "generating": SessionStatus.GENERATING,
        "awaiting_test_data_or_generate": SessionStatus.AWAITING_EARLY_TEST_DATA,
        "awaiting_review": SessionStatus.AWAITING_REVIEW,
        "improving": SessionStatus.IMPROVING,
        "generating_schema": SessionStatus.GENERATING_SCHEMA,
        "awaiting_test_data": SessionStatus.AWAITING_TEST_DATA,
        "generating_automation": SessionStatus.GENERATING_AUTOMATION,
        "awaiting_playwright_confirmation": SessionStatus.AWAITING_PLAYWRIGHT_CONFIRMATION,
        "ready_to_execute": SessionStatus.READY_TO_EXECUTE,
        "executing": SessionStatus.EXECUTING,
        "done": SessionStatus.DONE,
        "error": SessionStatus.ERROR,
    }
    return mapping.get(step or "", SessionStatus.IDLE)


# ---------------------------------------------------------------------------
# POST /chat/start
# ---------------------------------------------------------------------------

@router.post("/chat/start")
async def start_session(session_id: str, background_tasks: BackgroundTasks, g=Depends(get_graph)):
    """
    Kick off a new LangGraph thread for an already-uploaded session.
    Returns immediately with thread_id; the graph runs in a background task.
    Frontend should poll GET /chat/{thread_id}/status for progress.
    """
    thread_id = str(uuid.uuid4())
    config = _config(thread_id)

    upload_path = Path(settings.upload_dir) / session_id
    files = []
    
    files = [
        f for f in upload_path.glob("*")
        if f.is_file()
    ]
    if not files:
        raise HTTPException(status_code=404, detail="No uploaded file found for this session.")

    file_path = str(files[0])

    # Register the session in the persistent registry so it appears in session history.
    try:
        from app.services.session_registry import registry as _registry
        _registry.upsert_session(session_id=session_id, filename=files[0].name)
    except Exception:
        logger.warning("Failed to register session %s in registry", session_id)

    initial_state = {
        "messages": [HumanMessage(content="Start generating test cases from the uploaded file.")],
        "session_id": session_id,
        "thread_id": thread_id,
        "uploaded_file_path": file_path,
        "file_type": files[0].suffix.lstrip("."),
        "rag_context": "",
        "endpoints_summary": [],
        "api_metadata_map": {},
        "context_summary": "",
        "manual_test_cases": [],
        "test_suite_title": "",
        "human_approved": False,
        "human_feedback": None,
        "review_iteration": 0,
        "test_data_file_path": None,
        "test_data": [],
        "generated_test_file": None,
        "execution_status": None,
        "execution_log": [],
        "allure_report_url": None,
        "questionnaire_answers": {},
        "current_step": "parsing",
        "error_message": None,
    }

    # Run the graph in a background task so the HTTP response returns immediately.
    # The frontend polls /status to track progress as the graph advances.
    def _run_graph() -> None:
        try:
            # Stream 1: ingest_and_index runs, graph pauses before collect_questionnaire.
            list(g.stream(initial_state, config=config, stream_mode="values"))
            # Stream 2: auto-run collect_questionnaire (first pass) to generate questions
            # and set current_step="awaiting_questionnaire" without user action.
            mid = g.get_state(config)
            if mid and mid.values.get("current_step") != "error":
                list(g.stream(None, config=config, stream_mode="values"))
        except Exception:
            logger.exception("Background graph execution failed for thread %s", thread_id)

    background_tasks.add_task(_run_graph)

    return {
        "thread_id": thread_id,
        "session_id": session_id,
        "status": "parsing",
        "message": "File received. Parsing and indexing your spec — poll /status for progress updates.",
    }


# ---------------------------------------------------------------------------
# POST /chat/{thread_id}/message
# ---------------------------------------------------------------------------

@router.post("/chat/{thread_id}/message")
async def send_message(thread_id: str, body: ChatRequest, g=Depends(get_graph)):
    """
    Send a free-form message during the session (e.g. additional context, clarifications).
    The graph continues from its current interrupted state.
    """
    config = _config(thread_id)

    # Inject user message into state and resume
    g.update_state(config, {"messages": [HumanMessage(content=body.message)]})
    events = list(g.stream(None, config=config, stream_mode="values"))
    last_state = events[-1] if events else {}

    return {
        "thread_id": thread_id,
        "status": _map_status(last_state.get("current_step")),
        "message": _last_ai_message(last_state),
    }


# ---------------------------------------------------------------------------
# POST /chat/{thread_id}/review
# ---------------------------------------------------------------------------

@router.post("/chat/{thread_id}/review", response_model=ResumeResponse)
async def submit_review(
    thread_id: str,
    body: HumanReviewRequest,
    background_tasks: BackgroundTasks,
    g=Depends(get_graph),
):
    """
    Human submits review decision:
      - approved=True  → fully synchronous (schema generation is instant); schema message
                         returned directly in the response — no polling needed.
      - approved=False → LLM-intensive; returns immediately with "improving" status;
                         graph runs in a background task; frontend polls /status.
    """
    config = _config(thread_id)

    update: dict = {
        "human_approved": body.approved,
        "human_feedback": body.feedback,
        "current_step": "improving" if not body.approved else "awaiting_test_data",
    }
    # Apply any manual edits/deletions made by the user during review
    if body.approved and body.test_cases:
        update["manual_test_cases"] = body.test_cases
    g.update_state(config, update)

    if not body.approved:
        # Kick off improvement in the background — return immediately so the
        # frontend can start polling /status for progress.
        def _run_improvement() -> None:
            try:
                list(g.stream(None, config=config, stream_mode="values"))
            except Exception:
                logger.exception("Background improvement failed for thread %s", thread_id)

        background_tasks.add_task(_run_improvement)
        return ResumeResponse(
            thread_id=thread_id,
            message="Got it! Incorporating your feedback — this may take a minute. I'll update the test suite shortly.",
            status=_map_status("improving"),
        )

    # Approval path:
    # Save approved test cases to the persistent registry before the graph advances,
    # so they survive server restarts and are available in session history.
    try:
        from app.services.session_registry import registry as _registry
        _pre = g.get_state(config)
        if _pre:
            _registry.save_test_cases(
                session_id=_pre.values.get("session_id", ""),
                test_cases=_pre.values.get("manual_test_cases", []),
            )
    except Exception:
        logger.warning("Failed to save test cases to registry for thread %s", thread_id)

    #   Pass 1: resumes human_review → pauses before request_test_data
    list(g.stream(None, config=config, stream_mode="values"))
    #   Pass 2: runs request_test_data
    #     • Normal path (no early data): generates schema suggestion → pauses at request_test_data loop
    #     • Early-data path: finds test_data already set → sets current_step="generating_automation"
    #       → routes to generate_automated_tests → pauses there
    list(g.stream(None, config=config, stream_mode="values"))

    # Read the actual graph state after both passes to decide which path we're on.
    state_after = g.get_state(config)
    actual_step = state_after.values.get("current_step") if state_after else None

    if actual_step == "generating_automation":
        # Test data was uploaded early — graph is now paused at generate_automated_tests.
        # Kick off feature file generation in a background task, exactly like the
        # late-upload path does in documents.py, and return immediately so the frontend
        # can start polling for progress.
        def _run_feature_gen() -> None:
            try:
                from app.agents.nodes import generate_feature_files
                snap = g.get_state(config)
                generate_feature_files(g, config, snap.values)
            except Exception:
                logger.exception("Background feature generation failed for thread %s", thread_id)

        background_tasks.add_task(_run_feature_gen)
        return ResumeResponse(
            thread_id=thread_id,
            message=(
                "Test suite approved! Your test data is already loaded — "
                "generating feature files now. I'll update you when they're ready."
            ),
            status=_map_status("generating_automation"),
        )

    # Normal path — no early test data: return the schema suggestion message directly
    # so ReviewPanel can add it immediately without polling.
    schema_msg = _last_ai_message(state_after.values if state_after else {}) or (
        "Test suite approved! Upload a test data file using the 📎 button to generate automated tests."
    )
    return ResumeResponse(
        thread_id=thread_id,
        message=schema_msg,
        status=_map_status("awaiting_test_data"),
    )


# ---------------------------------------------------------------------------
# POST /chat/{thread_id}/questionnaire
# ---------------------------------------------------------------------------

@router.post("/chat/{thread_id}/questionnaire")
async def submit_questionnaire(
    thread_id: str,
    body: QuestionnaireSubmitRequest,
    g=Depends(get_graph),
):
    """
    Receive intake questionnaire answers from the frontend.
    Formats the answers into context_summary for later use by generate_test_cases,
    then pauses at awaiting_test_data_or_generate so the user can optionally upload
    test data before generation.  The graph is already paused at
    interrupt_before=["generate_test_cases"]; generation is kicked off by either
    POST /skip-early-data (no data) or POST /documents/test-data-early (with data).
    """
    from app.agents.nodes import _format_questionnaire_for_context
    config = _config(thread_id)

    # Read current questionnaire_questions so the formatter can label each answer
    current_state = g.get_state(config)
    vals = current_state.values if current_state else {}
    questions = vals.get("questionnaire_questions") or []
    existing_summary = vals.get("context_summary") or ""

    if body.answers:
        formatted = _format_questionnaire_for_context(body.answers, questions)
        combined_summary = (existing_summary + "\n\n" + formatted).strip()
    else:
        combined_summary = existing_summary

    g.update_state(config, {
        "questionnaire_answers": body.answers,
        "context_summary": combined_summary,
        "current_step": "awaiting_test_data_or_generate",
    })

    return {
        "thread_id": thread_id,
        "status": "awaiting_test_data_or_generate",
        "message": "Questionnaire received. Upload test data to enrich test cases, or skip to generate now.",
    }


# ---------------------------------------------------------------------------
# POST /chat/{thread_id}/skip-early-data
# ---------------------------------------------------------------------------

@router.post("/chat/{thread_id}/skip-early-data")
async def skip_early_test_data(thread_id: str, background_tasks: BackgroundTasks, g=Depends(get_graph)):
    """
    Resume the graph from the generate_test_cases interrupt without test data.
    Used when the user chooses to skip the optional early test data upload and
    proceed straight to test case generation.
    Runs in background; frontend polls /status for progress.
    """
    config = _config(thread_id)

    def _run_generate() -> None:
        try:
            list(g.stream(None, config=config, stream_mode="values"))
        except Exception:
            logger.exception("Background generate_test_cases failed for thread %s", thread_id)
            try:
                g.update_state(config, {
                    "current_step": "error",
                    "error_message": "Test case generation failed. Please start a new session.",
                })
            except Exception:
                pass

    background_tasks.add_task(_run_generate)
    return {
        "thread_id": thread_id,
        "status": "generating",
        "message": "Proceeding to test case generation — poll /status for progress.",
    }


# ---------------------------------------------------------------------------
# GET /chat/{thread_id}/status
# ---------------------------------------------------------------------------

@router.get("/chat/{thread_id}/status", response_model=SessionStatusResponse)
async def get_status(thread_id: str, session_id: str, g=Depends(get_graph)):
    config = _config(thread_id)
    state = g.get_state(config)
    values = state.values if state else {}

    return SessionStatusResponse(
        session_id=session_id,
        thread_id=thread_id,
        status=_map_status(values.get("current_step")),
        test_cases_count=len(values.get("manual_test_cases", [])),
        current_step=values.get("current_step", ""),
        last_message=_last_ai_message(values),
        questionnaire_questions=values.get("questionnaire_questions") or None,
    )


# ---------------------------------------------------------------------------
# GET /chat/{thread_id}/test-cases
# ---------------------------------------------------------------------------

@router.get("/chat/{thread_id}/test-cases")
async def get_test_cases(thread_id: str, g=Depends(get_graph)):
    config = _config(thread_id)
    state = g.get_state(config)
    values = state.values if state else {}

    test_cases = values.get("manual_test_cases", [])
    return {
        "thread_id": thread_id,
        "count": len(test_cases),
        "test_cases": test_cases,
    }


# ---------------------------------------------------------------------------
# GET /chat/{thread_id}/test-cases/export
# ---------------------------------------------------------------------------

@router.get("/chat/{thread_id}/test-cases/export")
async def export_test_cases(thread_id: str, g=Depends(get_graph)):
    """Download all manual test cases as a formatted Excel (.xlsx) file."""
    config = _config(thread_id)
    state = g.get_state(config)
    test_cases = (state.values if state else {}).get("manual_test_cases", [])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test Cases"

    headers = [
        "ID", "Title", "Module", "Test Type", "Priority",
        "Endpoint", "Preconditions", "Steps", "Expected Result",
        "Postconditions", "Notes",
    ]
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="4F81BD")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical="center")

    for tc in test_cases:
        preconditions = "\n".join(tc.get("preconditions") or [])
        steps_text = "\n".join(
            f"{s.get('step_number', i + 1)}. {s.get('action', '')}\n   \u2192 {s.get('expected_result', '')}"
            for i, s in enumerate(tc.get("steps") or [])
        )
        postconditions = "\n".join(tc.get("postconditions") or [])
        ws.append([
            tc.get("id", ""),
            tc.get("title", ""),
            tc.get("module", ""),
            tc.get("test_type", ""),
            tc.get("priority", ""),
            tc.get("endpoint", ""),
            preconditions,
            steps_text,
            tc.get("expected_result", ""),
            postconditions,
            tc.get("notes", ""),
        ])
        for cell in ws[ws.max_row]:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    col_widths = [10, 40, 20, 15, 12, 30, 40, 60, 50, 30, 30]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="test_cases_{thread_id[:8]}.xlsx"'},
    )


# ---------------------------------------------------------------------------
# GET /chat/{thread_id}/generated-test
# ---------------------------------------------------------------------------

@router.get("/chat/{thread_id}/generated-test")
async def get_generated_test(thread_id: str, g=Depends(get_graph)):
    """Return the content of the generated test file(s).

    For Gherkin output the state stores the path to the first .feature file.
    All sibling .feature files in the same directory are concatenated and returned
    so the frontend receives the full test suite in one response.
    """
    config = _config(thread_id)
    state = g.get_state(config)
    values = state.values if state else {}

    test_file = values.get("generated_test_file")
    if not test_file:
        raise HTTPException(status_code=404, detail="No generated test file found for this thread.")

    p = Path(test_file)
    if not p.exists():
        raise HTTPException(status_code=404, detail="Generated test file not found on disk.")

    try:
        if p.suffix == ".py":
            # Show the Gherkin feature files for human review;
            # test_generated.py is used for execution only.
            features_dir = p.parent / "features"
            if features_dir.exists():
                parts = [
                    f"# === {f.name} ===\n\n{f.read_text(encoding='utf-8')}"
                    for f in sorted(features_dir.glob("*.feature"))
                ]
                content = "\n\n".join(parts) if parts else p.read_text(encoding="utf-8")
            else:
                content = p.read_text(encoding="utf-8")
        elif p.suffix == ".feature" and p.parent.exists():
            # Legacy path — concatenate all sibling .feature files
            parts = []
            for feature_file in sorted(p.parent.glob("*.feature")):
                parts.append(f"# === {feature_file.name} ===\n\n{feature_file.read_text(encoding='utf-8')}")
            content = "\n\n".join(parts)
        else:
            content = p.read_text(encoding="utf-8")
    except OSError as exc:
        raise HTTPException(status_code=500, detail=f"Error reading generated file: {exc}")

    return {"thread_id": thread_id, "file_path": test_file, "content": content}


# ---------------------------------------------------------------------------
# GET /chat/{thread_id}/playwright-test
# ---------------------------------------------------------------------------

@router.get("/chat/{thread_id}/playwright-test")
async def get_playwright_test(thread_id: str, g=Depends(get_graph)):
    """Return the content of test_generated.py for display in the Playwright Tests tab."""
    config = _config(thread_id)
    state = g.get_state(config)
    values = state.values if state else {}

    test_file = values.get("generated_test_file")
    if not test_file:
        raise HTTPException(status_code=404, detail="No generated test file found for this thread.")

    p = Path(test_file)

    # If state points to test_generated.py directly, serve it
    if p.suffix == ".py":
        py_file = p
    else:
        # Legacy: state points to a .feature file — look for test_generated.py in parent
        py_file = p.parent.parent / "test_generated.py"

    if not py_file.exists():
        raise HTTPException(status_code=404, detail="Playwright test file not found on disk.")

    try:
        content = py_file.read_text(encoding="utf-8")
    except OSError as exc:
        raise HTTPException(status_code=500, detail=f"Error reading Playwright test file: {exc}")

    return {"thread_id": thread_id, "file_path": str(py_file), "content": content}


# ---------------------------------------------------------------------------
# POST /chat/{thread_id}/execute
# ---------------------------------------------------------------------------

@router.post("/chat/{thread_id}/execute")
async def execute_tests(thread_id: str, session_id: str, background_tasks: BackgroundTasks, g=Depends(get_graph)):
    """Kick off test execution in a background task; returns immediately."""
    config = _config(thread_id)
    state = g.get_state(config)
    values = state.values if state else {}

    test_file = values.get("generated_test_file")
    if not test_file:
        raise HTTPException(status_code=400, detail="Automated tests have not been generated yet.")

    # Signal "executing" immediately so the frontend polling sees the state change.
    g.update_state(
        config,
        {
            "current_step": "executing",
            "messages": [AIMessage(content="▶ Running automated tests — this may take a minute...")],
        },
    )

    def _run_tests() -> None:
        try:
            allure_results = os.path.join(settings.allure_results_dir, session_id)
            if os.path.exists(allure_results):
                shutil.rmtree(allure_results)
            os.makedirs(allure_results, exist_ok=True)
            cmd = [
                "python", "-m", "pytest",
                test_file,
                "--alluredir", allure_results,
                "-p", "no:allure_pytest_bdd",  # suppress plugin conflict when allure-pytest-bdd is also installed
                "-v",
                "--tb=short",
            ]
            result = subprocess.run(cmd, capture_output=True, text=True, timeout=300)

            passed = result.stdout.count(" PASSED")
            failed = result.stdout.count(" FAILED")
            status = "passed" if result.returncode == 0 else "failed"
            log_lines = (result.stdout + result.stderr).splitlines()

            # Try to generate the per-session Allure HTML report automatically
            report_url = f"/api/v1/report/view/{session_id}/index.html"
            try:
                generate_allure_report(session_id)
            except Exception:
                logger.warning("Allure report generation failed for thread %s — results still saved", thread_id)
                report_url = None

            summary = f"**Tests {status}** — {passed} passed, {failed} failed."
            if not report_url:
                summary += f"\n\n```\n{result.stdout[-2000:]}\n```"

            g.update_state(
                config,
                {
                    "execution_status": status,
                    "execution_log": log_lines,
                    "allure_report_url": report_url,
                    "current_step": "done",
                    "messages": [AIMessage(content=summary)],
                },
            )

            # Persist execution flag in the session registry so past-session viewer can show it
            try:
                registry.mark_has_execution(session_id, status)
            except Exception:
                logger.warning("mark_has_execution failed for session %s", session_id)

        except subprocess.TimeoutExpired:
            g.update_state(
                config,
                {
                    "execution_status": "error",
                    "current_step": "done",
                    "messages": [AIMessage(content="Test execution timed out after 5 minutes.")],
                },
            )
            try:
                registry.mark_has_execution(session_id, "error")
            except Exception:
                pass
        except Exception:
            logger.exception("Background test execution failed for thread %s", thread_id)
            g.update_state(
                config,
                {
                    "execution_status": "error",
                    "current_step": "done",
                    "messages": [AIMessage(content="Test execution encountered an unexpected error.")],
                },
            )
            try:
                registry.mark_has_execution(session_id, "error")
            except Exception:
                pass

    background_tasks.add_task(_run_tests)
    return {"thread_id": thread_id, "status": "executing", "message": "Tests started."}


# ---------------------------------------------------------------------------
# POST /chat/{thread_id}/generate-test-data
# ---------------------------------------------------------------------------

@router.post("/chat/{thread_id}/generate-test-data")
async def generate_test_data_from_spec(thread_id: str, n_rows: int = 6, g=Depends(get_graph)):
    """Generate realistic test data rows from the indexed API spec using the LLM."""
    from langchain_core.messages import SystemMessage, HumanMessage as LCHumanMessage
    from app.agents.prompts import TEST_DATA_GENERATION_PROMPT
    from app.agents.response_models import GeneratedTestDataOutput
    from app.rag.vector_store import similarity_search
    from app.services.llm_service import get_llm

    config = _config(thread_id)
    state = g.get_state(config)
    if not state or not state.values:
        raise HTTPException(status_code=400, detail="Session state not found.")

    values = state.values
    session_id: str = values.get("session_id", "")
    endpoints_summary: list[dict] = values.get("endpoints_summary", [])
    context_summary: str = values.get("context_summary", "")

    if not session_id:
        raise HTTPException(status_code=400, detail="No API spec indexed for this session.")

    # Build a readable summary of endpoints + their body fields
    ep_lines: list[str] = []
    for ep in endpoints_summary[:20]:
        method = ep.get("method", "")
        url = ep.get("url", ep.get("endpoint", ""))
        body_fields = ep.get("body_fields", [])
        fields_str = ", ".join(body_fields) if body_fields else "(no body)"
        ep_lines.append(f"  {method} {url}  —  body fields: {fields_str}")
    endpoints_text = "\n".join(ep_lines) if ep_lines else "No structured endpoints found."

    # RAG: pull payload/schema details from the indexed spec
    try:
        docs = similarity_search(
            session_id,
            "request payload body fields path parameters schema data types",
            k=10,
        )
        rag_context = "\n\n---\n\n".join(d.page_content for d in docs)
    except Exception:
        logger.warning("RAG retrieval failed for session %s; proceeding without context", session_id)
        rag_context = "(RAG context unavailable — using endpoint summary only)"

    context_section = f"Additional context:\n{context_summary}" if context_summary else ""

    n_rows = max(1, min(n_rows, 50))  # clamp 1–50
    system_prompt = TEST_DATA_GENERATION_PROMPT.format(
        rag_context=rag_context,
        endpoints_text=endpoints_text,
        context_section=context_section,
        n_rows=n_rows,
    )

    try:
        # temperature=0.5 intentional: diversity is the goal here; deterministic output at 0.0
        # produces near-identical rows that defeat the purpose of the test data.
        llm = get_llm(max_tokens=4000, temperature=0.5)
        result: GeneratedTestDataOutput = llm.with_structured_output(GeneratedTestDataOutput).invoke([
            SystemMessage(content=system_prompt),
            LCHumanMessage(content=f"Generate {n_rows} diverse, realistic test data rows for the API endpoints listed above."),
        ])
    except Exception:
        logger.exception("LLM test data generation failed for thread %s", thread_id)
        raise HTTPException(status_code=500, detail="Failed to generate test data. Please try again.")

    rows = [{f.key: f.value for f in row.fields} for row in (result.rows or [])]
    if not rows:
        raise HTTPException(status_code=500, detail="LLM returned no data rows. Please try again.")

    # Persist in graph state — same field used by uploaded test data
    g.update_state(config, {"test_data": rows})

    return {
        "thread_id": thread_id,
        "rows_generated": len(rows),
        "column_names": result.column_names,
        "data": rows,
        "message": f"Generated {len(rows)} test data rows from your API specification.",
    }


# ---------------------------------------------------------------------------
# GET /chat/{thread_id}/test-data/export
# ---------------------------------------------------------------------------

@router.get("/chat/{thread_id}/test-data/export")
async def export_test_data(thread_id: str, g=Depends(get_graph)):
    """Download the current test data (uploaded or AI-generated) as an Excel file."""
    config = _config(thread_id)
    state = g.get_state(config)
    values = state.values if state else {}
    test_data: list[dict] = values.get("test_data", [])

    if not test_data:
        raise HTTPException(status_code=404, detail="No test data found for this session.")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test Data"

    columns = list(test_data[0].keys())

    # Header row with styling
    ws.append(columns)
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row in test_data:
        ws.append([str(row.get(col, "")) for col in columns])

    # Auto-size columns
    for col_cells in ws.columns:
        width = max(len(str(cell.value or "")) for cell in col_cells)
        ws.column_dimensions[col_cells[0].column_letter].width = min(width + 4, 40)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=test_data_{thread_id[:8]}.xlsx"},
    )


# ---------------------------------------------------------------------------
# POST /chat/{thread_id}/confirm-playwright
# ---------------------------------------------------------------------------

@router.post("/chat/{thread_id}/confirm-playwright", response_model=ResumeResponse)
async def confirm_playwright(thread_id: str, background_tasks: BackgroundTasks, g=Depends(get_graph)):
    """User confirmed they want Playwright tests — kick off Phase 2 in the background."""
    config = _config(thread_id)
    g.update_state(config, {
        "current_step": "generating_automation",
        "messages": [AIMessage(content="Generating Playwright tests...")],
    })

    def _run() -> None:
        try:
            from app.agents.nodes import generate_playwright_tests
            snap = g.get_state(config)
            generate_playwright_tests(g, config, snap.values)
        except Exception:
            logger.exception("Background Playwright generation failed for thread %s", thread_id)

    background_tasks.add_task(_run)
    return ResumeResponse(
        thread_id=thread_id,
        message="Generating Playwright tests...",
        status=_map_status("generating_automation"),
    )


# ---------------------------------------------------------------------------
# POST /chat/{thread_id}/skip-playwright
# ---------------------------------------------------------------------------

@router.post("/chat/{thread_id}/skip-playwright", response_model=ResumeResponse)
async def skip_playwright(thread_id: str, g=Depends(get_graph)):
    """User chose to skip Playwright generation — go straight to ready_to_execute."""
    config = _config(thread_id)
    g.update_state(config, {
        "current_step": "ready_to_execute",
        "messages": [AIMessage(content=(
            "Playwright generation skipped. Switch to the **Feature Files** tab to review, "
            "or click **Run Tests** if you have an existing test suite."
        ))],
    })
    return ResumeResponse(
        thread_id=thread_id,
        message="Skipped Playwright generation.",
        status=_map_status("ready_to_execute"),
    )


# ---------------------------------------------------------------------------
# POST /chat/{thread_id}/skip-load-test
# ---------------------------------------------------------------------------

@router.post("/chat/{thread_id}/skip-load-test")
async def skip_load_test(thread_id: str, g=Depends(get_graph)):
    """Advance from awaiting_load_test_config → ready_to_execute (skip or done creating)."""
    config = _config(thread_id)
    g.update_state(config, {
        "current_step": "ready_to_execute",
        "messages": [AIMessage(content="Load test configuration complete. Ready to execute tests.")],
    })
    return {"status": "ok", "thread_id": thread_id}


# ---------------------------------------------------------------------------
# POST /chat/{thread_id}/generate-load-test
# ---------------------------------------------------------------------------

class LoadTestGenerateRequest(BaseModel):
    name: str = "Load Test"
    selected_endpoints: list[str]
    vus: int = 10
    duration: str = "2m"
    ramp_up: str = "30s"
    ramp_down: str = "30s"
    p95_ms: int = 500
    p99_ms: int = 1000
    error_rate_pct: float = 1.0


def _strip_code_fences(text: str) -> str:
    """Remove markdown code fences (```javascript ... ``` or ``` ... ```) from LLM output."""
    text = re.sub(r"^```[a-zA-Z]*\n?", "", text.strip())
    text = re.sub(r"\n?```$", "", text.strip())
    return text.strip()


@router.post("/chat/{thread_id}/generate-load-test")
async def generate_load_test(
    thread_id: str,
    body: LoadTestGenerateRequest,
    g=Depends(get_graph),
):
    """
    Generate a single load test script for the selected endpoints.
    Synchronous LLM call (runs in executor); returns script content directly.
    """
    from app.agents.prompts import GENERATE_LOAD_TEST_PROMPT
    from app.services.llm_service import get_llm

    if not body.selected_endpoints:
        raise HTTPException(status_code=400, detail="No endpoints selected.")

    config = _config(thread_id)
    state = g.get_state(config)
    if not state:
        raise HTTPException(status_code=404, detail="Thread not found.")
    values = state.values

    session_id = values.get("session_id", "")
    api_metadata_map: dict = values.get("api_metadata_map") or {}
    test_data: list = values.get("test_data") or []
    questionnaire_answers: dict = values.get("questionnaire_answers") or {}
    existing_load_tests: list = values.get("load_tests") or []

    # Build endpoint context
    _BUDGET = 1200
    ep_sections = []
    for ep_key in body.selected_endpoints:
        meta = api_metadata_map.get(ep_key) or {}
        content = meta.get("content", ep_key)
        ep_sections.append(f"### {ep_key}\n{content[:_BUDGET]}")
    api_metadata_str = "\n\n".join(ep_sections) if ep_sections else "(no metadata available)"

    endpoints_ordered = "\n".join(f"{i+1}. {ep}" for i, ep in enumerate(body.selected_endpoints))

    test_data_sample = ""
    if test_data:
        cols = list(test_data[0].keys()) if test_data else []
        rows = test_data[:3]
        sample_lines = [", ".join(f"{k}={v}" for k, v in row.items()) for row in rows]
        test_data_sample = f"Columns: {cols}\nSample rows:\n" + "\n".join(sample_lines)
    else:
        test_data_sample = "(no test data provided)"

    s1 = questionnaire_answers.get("s1", {})
    auth_type = s1.get("auth_type", "Bearer Token")
    base_url = s1.get("base_url", "http://localhost:8000")
    auth_info = f"Auth type: {auth_type}\nBase URL: {base_url}"

    prompt = GENERATE_LOAD_TEST_PROMPT.format(
        load_test_name=body.name,
        endpoints_ordered=endpoints_ordered,
        api_metadata=api_metadata_str,
        test_data_sample=test_data_sample,
        auth_info=auth_info,
        vus=body.vus,
        duration=body.duration,
        ramp_up=body.ramp_up,
        ramp_down=body.ramp_down,
        p95_ms=body.p95_ms,
        p99_ms=body.p99_ms,
        error_rate_pct=body.error_rate_pct,
        error_rate_pct_decimal=body.error_rate_pct / 100,
    )

    def _run_llm() -> str:
        llm = get_llm(max_tokens=6000, temperature=0.2)
        result = llm.invoke(prompt)
        raw = result.content
        text = raw if isinstance(raw, str) else " ".join(str(p) for p in raw)
        return _strip_code_fences(text)

    script_content = await asyncio.get_event_loop().run_in_executor(None, _run_llm)

    # Persist to disk
    load_tests_dir = Path(settings.generated_tests_dir) / session_id / "load_tests"
    load_tests_dir.mkdir(parents=True, exist_ok=True)

    lt_index = len(existing_load_tests) + 1
    lt_id = f"lt_{lt_index}"
    file_name = f"load_test_{lt_index}.js"
    file_path = str(load_tests_dir / file_name)
    Path(file_path).write_text(script_content, encoding="utf-8")

    # Build manifest entry
    new_entry = {
        "id": lt_id,
        "name": body.name,
        "endpoints": body.selected_endpoints,
        "file": file_name,
        "file_path": file_path,
        "vus": body.vus,
        "duration": body.duration,
        "ramp_up": body.ramp_up,
        "ramp_down": body.ramp_down,
        "p95_ms": body.p95_ms,
        "p99_ms": body.p99_ms,
        "error_rate_pct": body.error_rate_pct,
    }

    # Update manifest.json
    manifest_path = load_tests_dir / "manifest.json"
    existing_manifest = {"load_tests": list(existing_load_tests)}
    if manifest_path.exists():
        try:
            existing_manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
        except Exception:
            pass
    existing_manifest["load_tests"].append({k: v for k, v in new_entry.items() if k != "file_path"})
    manifest_path.write_text(json.dumps(existing_manifest, indent=2), encoding="utf-8")

    # Update LangGraph state
    updated_load_tests = list(existing_load_tests) + [new_entry]
    g.update_state(config, {"load_tests": updated_load_tests})

    # Mark registry
    try:
        registry.mark_has_load_tests(session_id)
    except Exception:
        logger.warning("mark_has_load_tests failed for session %s", session_id)

    return {
        "id": lt_id,
        "name": body.name,
        "endpoints": body.selected_endpoints,
        "file_path": file_path,
        "content": script_content,
        "vus": body.vus,
        "duration": body.duration,
    }


# ---------------------------------------------------------------------------
# GET /chat/{thread_id}/load-tests
# ---------------------------------------------------------------------------

@router.get("/chat/{thread_id}/load-tests")
async def get_load_tests(thread_id: str, g=Depends(get_graph)):
    """Return all generated load test scripts for this session."""
    config = _config(thread_id)
    state = g.get_state(config)
    if not state:
        raise HTTPException(status_code=404, detail="Thread not found.")

    session_id = state.values.get("session_id", "")
    load_tests_dir = Path(settings.generated_tests_dir) / session_id / "load_tests"

    if not load_tests_dir.exists():
        return {"load_tests": []}

    manifest_path = load_tests_dir / "manifest.json"
    if not manifest_path.exists():
        return {"load_tests": []}

    try:
        manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    except Exception:
        return {"load_tests": []}

    result = []
    for entry in manifest.get("load_tests", []):
        js_path = load_tests_dir / entry["file"]
        content = ""
        if js_path.exists():
            try:
                content = js_path.read_text(encoding="utf-8")
            except Exception:
                pass
        result.append({
            "id": entry.get("id"),
            "name": entry.get("name"),
            "endpoints": entry.get("endpoints", []),
            "file_path": str(js_path),
            "content": content,
            "vus": entry.get("vus", 10),
            "duration": entry.get("duration", "2m"),
            "ramp_up": entry.get("ramp_up", "30s"),
            "ramp_down": entry.get("ramp_down", "30s"),
            "p95_ms": entry.get("p95_ms", 500),
            "p99_ms": entry.get("p99_ms", 1000),
            "error_rate_pct": entry.get("error_rate_pct", 1.0),
        })

    return {"load_tests": result}
